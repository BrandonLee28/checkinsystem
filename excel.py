import sqlite3
from openpyxl import Workbook
import io

def export_cursor_to_xlsx(cursor):
    # Create a new workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Fetch all the rows from the cursor
    rows = cursor.fetchall()

    # Write the column names as the first row
    column_names = [description[0] for description in cursor.description]
    for col_idx, column_name in enumerate(column_names, 1):
        worksheet.cell(row=1, column=col_idx, value=column_name)

    # Write the data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Save the workbook to a BytesIO buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer

# Example usage
